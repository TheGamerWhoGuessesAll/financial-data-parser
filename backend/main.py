import io
import os
import json
import uuid
import asyncio
import re
import pandas as pd
import pdfplumber
from fastapi.responses import RedirectResponse
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks, Depends, Request
from fastapi.responses import StreamingResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from datetime import datetime, timedelta
from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Boolean, text
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from passlib.context import CryptContext
import jwt
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from fastapi.security import OAuth2PasswordBearer
from fastapi import Depends, Request

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart3D, Reference
from openpyxl.chart.label import DataLabelList


POLICY_THRESHOLDS = {
    "food": 50,
    "meals": 50,
    "fast food": 50,
    "restaurants": 50,
    "coffee": 20,
    "travel": 1000,
    "airlines": 1000,
    "hotel": 500,
    "software": 200,
    "office supplies": 100,
    "fuel": 100,
    "gas station": 100,
    "entertainment": 0,
}

JSON_PATTERN = re.compile(r'\{.*\}', re.DOTALL)

def apply_policy_rules(category, amount, raw_text="", is_suspicious=False, suspicion_reason=""):
    if is_suspicious:
        return f"Requires Review: AI Flag - {suspicion_reason}"
        
    raw_lower = str(raw_text).lower()
    restricted_terms = ["fraud", "suspicious", "unauthorized", "scam", "stolen", "theft"]
    if any(term in raw_lower for term in restricted_terms):
        return "Requires Review: Flagged Keywords"
        
    if amount is None:
        return "Requires Review: Invalid amount"
        
    if isinstance(amount, str):
        amount = amount.translate(str.maketrans('', '', '$,')).strip()
        
    try:
        amount = float(amount)
    except (ValueError, TypeError):
        return "Requires Review: Invalid amount"
        
    category_str = str(category).strip() if category is not None else "Other"
    threshold = POLICY_THRESHOLDS.get(category_str.lower(), 500)
    
    if amount > threshold:
        return f"Requires Review: Amount {amount} exceeds {category_str} policy limit of {threshold}"
    return "Compliant"

app = FastAPI(title="Financial Data Parser")





app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://financial-data-parser.onrender.com",
        "https://finparse.dev"
    ],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)


# Database Setup
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./test.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False} if "sqlite" in DATABASE_URL else {})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


class ProcessedEvent(Base):
    __tablename__ = "processed_events"
    id = Column(String, primary_key=True)

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String, unique=True, index=True)
    hashed_password = Column(String)
    failed_login_attempts = Column(Integer, default=0)
    locked_until = Column(DateTime, default=None)
    lockout_count = Column(Integer, default=0)
    reset_token = Column(String, default=None, index=True)
    reset_token_expiry = Column(DateTime, default=None)
    rows_processed_this_month = Column(Integer, default=0)
    last_reset_date = Column(DateTime, default=None)
    subscription_tier = Column(String, default='free')
    is_verified = Column(Boolean, default=False)

Base.metadata.create_all(bind=engine)

from sqlalchemy import text
try:
    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE users ADD COLUMN is_verified BOOLEAN DEFAULT FALSE"))
        conn.commit()
except Exception:
    pass

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Auth Setup
SECRET_KEY = os.environ["SECRET_KEY"]
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7 # 7 days

import bcrypt
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

def verify_password(plain_password, hashed_password):
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def get_password_hash(password):
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt(rounds=14)).decode('utf-8')

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=401,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
        
    user = db.query(User).filter(User.email == email).first()
    if user is None:
        raise credentials_exception
    return user

# Pydantic Models
class UserCreate(BaseModel):
    email: str
    password: str

class UserLogin(BaseModel):
    email: str
    password: str

# Rate Limiter Setup
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

@app.post("/register")
@limiter.limit("5/minute")
def signup(request: Request, user: UserCreate, db: Session = Depends(get_db)):
    email_lower = user.email.lower()
    db_user = db.query(User).filter(User.email == email_lower).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    hashed_password = get_password_hash(user.password)
    
    # Auto-VIP for the testing account
    tier = 'unlimited' if email_lower == 'mg.shuanchi@gmail.com' else 'free'
    
    new_user = User(email=email_lower, hashed_password=hashed_password, is_verified=False, subscription_tier=tier)
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    token = create_access_token(data={"sub": new_user.email})
    verify_link = f"https://financial-data-parser.onrender.com/verify.html?token={token}"
    
    resend_api_key = os.getenv("RESEND_API_KEY")
    if resend_api_key:
        try:
            req_data = json.dumps({
                "from": "onboarding@resend.dev",
                "to": [user.email],
                "subject": "Verify Your FinParse Account",
                "html": f"<p>Welcome to FinParse!</p><p>Please click the link below to verify your email address:</p><p><a href='{verify_link}'>{verify_link}</a></p>"
            }).encode('utf-8')
            
            headers = {
                "Authorization": f"Bearer {resend_api_key}",
                "Content-Type": "application/json"
            }
            
            http_req = urllib.request.Request("https://api.resend.com/emails", data=req_data, headers=headers, method="POST")
            urllib.request.urlopen(http_req)
        except Exception as e:
            print("Failed to send verification email:", e)
            
    return {"message": "Please check your email to verify your account before logging in."}

class VerifyRequest(BaseModel):
    token: str

@app.post("/verify")
def verify_account(req: VerifyRequest, db: Session = Depends(get_db)):
    try:
        payload = jwt.decode(req.token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise HTTPException(status_code=400, detail="Invalid token")
    except jwt.PyJWTError:
        raise HTTPException(status_code=400, detail="Invalid or expired token")
        
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    user.is_verified = True
    db.commit()
    return {"message": "Account successfully verified!"}

@app.post("/login")
@limiter.limit("5/minute")
def login(request: Request, user: UserLogin, db: Session = Depends(get_db)):
    email_lower = user.email.lower()
    db_user = db.query(User).filter(User.email == email_lower).first()
    
    if not db_user:
        raise HTTPException(status_code=401, detail="Incorrect email or password")
        
    if not db_user.is_verified:
        raise HTTPException(status_code=403, detail="Please check your email and verify your account first.")
        
    if db_user.locked_until and db_user.locked_until > datetime.utcnow():
        raise HTTPException(status_code=403, detail=f"Account locked until {db_user.locked_until.strftime('%H:%M:%S')} UTC due to too many failed attempts.")
        
    if not verify_password(user.password, db_user.hashed_password):
        db_user.failed_login_attempts += 1
        if db_user.failed_login_attempts >= 5:
            db_user.lockout_count += 1
            lockout_mins = min(15 * (2 ** (db_user.lockout_count - 1)), 24 * 60)
            db_user.locked_until = datetime.utcnow() + timedelta(minutes=lockout_mins)
            db.commit()
            raise HTTPException(status_code=403, detail=f"Account locked for {lockout_mins} minutes due to too many failed attempts.")
            
        db.commit()
        raise HTTPException(status_code=401, detail="Incorrect email or password")
        
    # Correct password - reset counters
    db_user.failed_login_attempts = 0
    db_user.lockout_count = 0
    db_user.locked_until = None
    
    # Auto-upgrade testing account if it wasn't already upgraded
    if email_lower == 'mg.shuanchi@gmail.com' and db_user.subscription_tier != 'unlimited':
        db_user.subscription_tier = 'unlimited'
        
    db.commit()
    
    access_token = create_access_token(data={"sub": db_user.email})
    return {"access_token": access_token, "token_type": "bearer"}

TASK_STORE = {}


@app.post("/upload")
async def upload_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    is_pdf = file.filename and file.filename.lower().endswith('.pdf')
    is_csv = file.filename and file.filename.lower().endswith('.csv')
    
    if not (is_pdf or is_csv):
        raise HTTPException(status_code=400, detail="Only .csv and .pdf files are supported")

    # Limit checking will happen in the background task to avoid blocking the HTTP response
    # Read the uploaded file
    contents = await file.read()
    
    task_id = str(uuid.uuid4())
    TASK_STORE[task_id] = {
        "status": "processing",
        "progress": 0,
        "message": "Initializing task...",
        "result": None,
        "filename": file.filename or "data.csv",
        "user_email": current_user.email
    }
    
    background_tasks.add_task(process_file_task, task_id, contents, is_csv, is_pdf, current_user.id)
    return {"task_id": task_id}

@app.get("/status/{task_id}")
async def get_status(task_id: str, current_user: User = Depends(get_current_user)):
    if task_id not in TASK_STORE or TASK_STORE[task_id].get("user_email") != current_user.email:
        raise HTTPException(status_code=404, detail="Task not found")
    return {
        "status": TASK_STORE[task_id]["status"],
        "progress": TASK_STORE[task_id]["progress"],
        "message": TASK_STORE[task_id]["message"]
    }

@app.get("/download/{task_id}")
async def download_file(task_id: str, format: str = "excel", current_user: User = Depends(get_current_user)):
    if task_id not in TASK_STORE or TASK_STORE[task_id].get("user_email") != current_user.email:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task_data = TASK_STORE[task_id]
    if task_data["status"] != "completed":
        raise HTTPException(status_code=400, detail="Task not completed yet")
        
    buffer = io.BytesIO(task_data["result"])
    buffer.seek(0)
    
    original_filename = task_data["filename"]
    base_name = original_filename.rsplit('.', 1)[0]
    excel_filename = f"{base_name}-Examination.xlsx"
    
    del TASK_STORE[task_id] # Clean up RAM
    
    if format == "csv" and current_user.package != "basic":
        import pandas as pd
        df = pd.read_excel(buffer, sheet_name="Anomaly Report")
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_bytes = csv_buffer.getvalue().encode('utf-8')
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={base_name}-Anomalies.csv"}
        )
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={excel_filename}"}
    )

async def process_file_task(task_id: str, contents: bytes, is_csv: bool, is_pdf: bool, user_id: int):
    try:
        global_warnings = []
        TASK_STORE[task_id]["message"] = "Extracting data..."
        TASK_STORE[task_id]["progress"] = 10
        
        if is_csv:
            df = pd.read_csv(io.BytesIO(contents))
        else:
            all_rows = []
            with pdfplumber.open(io.BytesIO(contents)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        for row in table:
                            if any(cell and str(cell).strip() for cell in row):
                                all_rows.append(row)
            if not all_rows:
                raise ValueError("No tabular data found in PDF")
            df = pd.DataFrame(all_rows[1:], columns=all_rows[0])
            for col in df.columns:
                if pd.api.types.is_string_dtype(df[col]):
                    df[col] = df[col].astype(str).str.strip()
                    try:
                        clean_col = df[col].str.replace('$', '', regex=False).str.replace(',', '', regex=False).str.replace(' ', '', regex=False)
                        num_col = pd.to_numeric(clean_col)
                        if num_col.notna().sum() > (len(df) * 0.3):
                            df[col] = num_col
                    except:
                        pass
                        
        import re
        
        df_str = df.astype(object).fillna("").astype(str)
        has_newline = df_str.apply(lambda x: x.str.count(r'[\n\r]') > 4).any(axis=1)
        date_pattern = r'\b\d{2,4}[-/]\d{1,2}[-/]\d{2,4}\b'
        has_multiple_dates = df_str.apply(lambda x: x.str.count(date_pattern) >= 4).any(axis=1)
        is_all_empty = df_str.apply(lambda x: x.str.strip() == "").all(axis=1)
        has_html_tags = df_str.apply(lambda x: x.str.contains(r'<[^>]+>', na=False, regex=True)).any(axis=1)
        is_header_repeat = df_str.apply(lambda x: x.str.match(r'(?i)^\s*(date|amount|description|merchant)\s*$')).any(axis=1)
        has_absurd_length = df_str.apply(lambda x: x.str.len() > 500).any(axis=1)
        
        gibberish_pattern = r'^([^a-zA-Z0-9]*[^a-zA-Z0-9\s][^a-zA-Z0-9]*|[^ \d]{20,})$'
        is_gibberish = df_str.apply(lambda x: x.str.match(gibberish_pattern)).any(axis=1)
        
        malformed_mask = (has_newline | has_multiple_dates | is_all_empty | 
                          has_html_tags | is_header_repeat | has_absurd_length | is_gibberish)
        
        num_malformed = int(malformed_mask.sum())
        
        if num_malformed > 0:
            df = df[~malformed_mask].reset_index(drop=True)
            global_warnings.append(f"{num_malformed} malformed rows were skipped (failed advanced validation: empty, html, gibberish, etc.).")
            
        # --- ROW COUNTING & LIMIT CHECKING ---
        effective_rows = sum(max(1, len(str(row)) // 200) for _, row in df.iterrows())
        
        db = SessionLocal()
        try:
            user = db.query(User).filter(User.id == user_id).first()
            if not user:
                raise ValueError("User not found")
                
            # Check for monthly reset
            now = datetime.now()
            if not user.last_reset_date or user.last_reset_date.month != now.month or user.last_reset_date.year != now.year:
                user.rows_processed_this_month = 0
                user.last_reset_date = now
                
            # Determine Tier Limit
            tier = user.subscription_tier or 'free'
            if tier == 'free':
                limit = 100
            elif tier == 'budget':
                limit = 1000
            elif tier == 'pro':
                limit = 5000
            else:
                limit = float('inf')
                
            if user.rows_processed_this_month + effective_rows > limit:
                raise ValueError(f"Limit Exceeded: This file contains {effective_rows} rows, but you only have {limit - user.rows_processed_this_month} rows remaining in your {tier.capitalize()} plan this month.")
                
            user.rows_processed_this_month += effective_rows
            db.commit()
        finally:
            db.close()
        # ------------------------------------
        
        TASK_STORE[task_id]["message"] = "Cleaning data & running Math Engine..."
        TASK_STORE[task_id]["progress"] = 30
        
        # Dynamically scan contents for Date/Time format
        date_col = None
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                date_col = col
                break
            
            if pd.api.types.is_string_dtype(df[col]) or pd.api.types.is_object_dtype(df[col]):
                sample_idx = df[col].first_valid_index()
                if sample_idx is not None:
                    sample_val = str(df[col].loc[sample_idx])
                    if any(c in sample_val for c in ['-', '/', ':']):
                        try:
                            pd.to_datetime(sample_val)
                            parsed_col = pd.to_datetime(df[col], errors='coerce')
                            if parsed_col.notna().sum() > (df[col].notna().sum() * 0.5):
                                date_col = col
                                df[date_col] = parsed_col
                                break
                        except Exception:
                            pass
                
        if date_col:
            if not pd.api.types.is_datetime64_any_dtype(df[date_col]):
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                
            df = df.sort_values(by=date_col).reset_index(drop=True)
            df[date_col] = df[date_col].dt.strftime('%Y-%m-%d %H:%M:%S').fillna(df[date_col])
            
        # Rapid Succession Detection
        rapid_warnings = 0
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                vals = df[col].values
                streak = 1
                for i in range(1, len(vals)):
                    prev = vals[i-1]
                    curr = vals[i]
                    if pd.isna(prev) or pd.isna(curr) or prev == 0:
                        if streak >= 3:
                            rapid_warnings += 1
                        streak = 1
                        continue
                    diff_pct = abs(curr - prev) / abs(prev)
                    if diff_pct <= 0.05:
                        streak += 1
                    else:
                        if streak >= 3:
                            rapid_warnings += 1
                        streak = 1
                if streak >= 3:
                    rapid_warnings += 1


        # AI Contextual Assessment
        ai_assessments = ["AI Pending"] * len(df)
        
        TASK_STORE[task_id]["message"] = "Connecting to Gemini AI..."
        TASK_STORE[task_id]["progress"] = 40
            
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            ai_assessments = ["API Key Missing"] * len(df)
        else:
            try:
                from google import genai
                from google.genai import types
                
                client = genai.Client(api_key=api_key)
                desc_col = None
                amt_col = None
                for col in df.columns:
                    if pd.api.types.is_numeric_dtype(df[col]) and amt_col is None:
                        amt_col = col
                for col in df.columns:
                    if (pd.api.types.is_string_dtype(df[col]) or pd.api.types.is_object_dtype(df[col])) and date_col != col:
                        if 'desc' in str(col).lower() or 'name' in str(col).lower() or 'merchant' in str(col).lower():
                            desc_col = col
                            break
                if desc_col is None:
                    for col in df.columns:
                        if (pd.api.types.is_string_dtype(df[col]) or pd.api.types.is_object_dtype(df[col])) and date_col != col:
                            desc_col = col
                            break
                
                if desc_col and amt_col:
                    batch_data = []
                    for idx, row in df.iterrows():
                        batch_data.append({
                            "id": idx,
                            "desc": str(row[desc_col]),
                            "amount": float(row[amt_col]) if not pd.isna(row[amt_col]) else 0.0
                        })
                    
                    batch_size = 50
                    chunks = [batch_data[i:i+batch_size] for i in range(0, len(batch_data), batch_size)]
                    total_chunks = len(chunks)
                    chunks_completed = 0
                    
                    TASK_STORE[task_id]["message"] = f"AI Analyzing {len(df)} transactions..."
                    TASK_STORE[task_id]["progress"] = 50
                    
                    # Limit concurrency to 1 to completely avoid Free Tier burst rate limits
                    semaphore = asyncio.Semaphore(1)
                    
                    async def process_chunk(chunk):
                        nonlocal chunks_completed
                        async with semaphore:
                            prompt = f"""
                            You are a data extraction assistant. Analyze the following list of transactions.
                            Extract the Vendor Name, Amount, Category, is_suspicious, and suspicion_reason for each transaction.
                            The category MUST be one of the following ~50 categories:
                            "Advertising", "Airlines", "Automotive", "Bakery", "Bank Fees", "Bookstores", "Car Rental",
                            "Charity", "Clothing", "Coffee", "Consulting", "Convenience Store", "Cosmetics", "Coworking",
                            "Delivery", "Department Store", "Education", "Electronics", "Entertainment", "Fast Food", 
                            "Fitness", "Food", "Fuel", "Furniture", "Gas Station", "Groceries", "Hardware", "Health", 
                            "Home Improvement", "Hotel", "Insurance", "Internet", "Legal", "Logistics", "Maintenance", 
                            "Meals", "Medical", "Office Supplies", "Online Retail", "Parking", "Personal Care", "Pets", 
                            "Pharmacy", "Postage", "Printing", "Public Transit", "Real Estate", "Restaurants", "Rideshare",
                            "Software", "Sporting Goods", "Streaming", "Subscriptions", "Supermarket", "Taxis", "Telecommunications", 
                            "Tolls", "Travel", "Utilities", "Wholesale", "Other".
                            
                            Return a JSON object mapping the transaction 'id' to an object containing:
                            - vendor: The extracted vendor name (string)
                            - amount: The numeric amount (number)
                            - category: The best-matching category from the list above (string)
                            - is_suspicious: evaluating if contextually it looks like fraud/unauthorized (boolean)
                            - suspicion_reason: explaining why if true, otherwise empty (string)
                            
                            Evaluate the context, vendor name, and amount. Is this likely fraud, unauthorized, or highly suspicious? Set is_suspicious to true if so.
                            
                            Transactions:
                            {json.dumps(chunk)}
                            """
                            
                            last_error = "Unknown error"
                            for attempt in range(3):
                                try:
                                    # Use the asynchronous client
                                    response = await client.aio.models.generate_content(
                                        model='gemini-2.5-flash',
                                        contents=prompt,
                                        config=types.GenerateContentConfig(
                                            response_mime_type="application/json",
                                        ),
                                    )
                                    text = response.text.strip()
                                    if text.startswith("```json"):
                                        text = text[7:]
                                    elif text.startswith("```"):
                                        text = text[3:]
                                    if text.endswith("```"):
                                        text = text[:-3]
                                    text = text.strip()
                                        
                                    try:
                                        parsed = json.loads(text)
                                    except json.JSONDecodeError:
                                        match = JSON_PATTERN.search(text)
                                        if match:
                                            parsed = json.loads(match.group(0))
                                        else:
                                            raise
                                            
                                    if parsed is None:
                                        parsed = {}
                                        
                                    chunks_completed += 1
                                    progress_pct = 50 + int((chunks_completed / total_chunks) * 40)
                                    TASK_STORE[task_id]["progress"] = progress_pct
                                    TASK_STORE[task_id]["message"] = f"AI Analyzed {min(chunks_completed * batch_size, len(df))} of {len(df)} transactions..."
                                        
                                    return parsed
                                except Exception as e:
                                    last_error = str(e)
                                    if hasattr(e, 'message'):
                                        last_error = e.message
                                    if attempt < 2:
                                        await asyncio.sleep(2 ** attempt)
                                    else:
                                        chunks_completed += 1
                                        progress_pct = 50 + int((chunks_completed / total_chunks) * 40)
                                        TASK_STORE[task_id]["progress"] = progress_pct
                                        TASK_STORE[task_id]["message"] = f"AI Analyzed {min(chunks_completed * batch_size, len(df))} of {len(df)} transactions..."
                                        return {str(item["id"]): f"AI Error: {last_error}" for item in chunk}
                            
                            chunks_completed += 1
                            progress_pct = 50 + int((chunks_completed / total_chunks) * 40)
                            TASK_STORE[task_id]["progress"] = progress_pct
                            TASK_STORE[task_id]["message"] = f"AI Analyzed {min(chunks_completed * batch_size, len(df))} of {len(df)} transactions..."
                            return {str(item["id"]): f"AI Error: {last_error}" for item in chunk}
                    
                    tasks = [process_chunk(chunk) for chunk in chunks]
                    results = await asyncio.gather(*tasks)
                    
                    for assessments in results:
                        for idx_str, data in assessments.items():
                            idx_int = int(idx_str)
                            if 0 <= idx_int < len(ai_assessments):
                                if isinstance(data, dict):
                                    is_susp_raw = data.get("is_suspicious", False)
                                    is_susp = is_susp_raw if isinstance(is_susp_raw, bool) else str(is_susp_raw).strip().lower() == "true"
                                    susp_reason = data.get("suspicion_reason", "")
                                    
                                    if is_susp:
                                        ai_assessments[idx_int] = apply_policy_rules(None, None, "", is_susp, susp_reason)
                                    else:
                                        cat = data.get("category", "Other")
                                        amt = data.get("amount", 0.0)
                                        raw_text = " ".join([str(val) for val in df.values[idx_int]])
                                        ai_assessments[idx_int] = apply_policy_rules(cat, amt, raw_text, is_susp, susp_reason)
                                else:
                                    ai_assessments[idx_int] = str(data)
                else:
                    ai_assessments = ["Missing Amt/Desc Col"] * len(df)
            except Exception as e:
                err_msg = str(e)
                if hasattr(e, 'message'):
                    err_msg = e.message
                ai_assessments = [f"AI Error (Outer): {type(e).__name__} - {err_msg}"] * len(df)
                
        df['Policy Assessment'] = ai_assessments
        
        # Terminology Update: Replace raw occurrences of 'Fraud' and 'Suspicious' in the final output
        df = df.replace(to_replace=r'(?i)fraud', value='Requires Review', regex=True)
        df = df.replace(to_replace=r'(?i)suspicious', value='Requires Review', regex=True)

        TASK_STORE[task_id]["message"] = "Generating Excel Report..."
        TASK_STORE[task_id]["progress"] = 80

        # Create an in-memory buffer for the Excel file
        buffer = io.BytesIO()

        # Formatting
        yellow_fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
        yellow_font = Font(color="FF9C6500", bold=True)
        orange_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        orange_font = Font(color="FF9C0006", bold=True)
        red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
        red_font = Font(color="FF990000", bold=True)
        severe_fill = PatternFill(start_color="FF990000", end_color="FF990000", fill_type="solid")
        severe_font = Font(color="FFFFFFFF", bold=True)
        purple_fill = PatternFill(start_color="FFCCCCFF", end_color="FFCCCCFF", fill_type="solid")
        purple_font = Font(color="FF333399", bold=True)

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Anomaly Report')
            worksheet = writer.sheets['Anomaly Report']
            
            worksheet.freeze_panes = "A2"
            max_col_letter = get_column_letter(len(df.columns))
            max_row = len(df) + 1
            worksheet.auto_filter.ref = f"A1:{max_col_letter}{max_row}"
            
            ai_col_idx = df.columns.get_loc('Policy Assessment') + 1
            
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 50)
                
            numeric_cols = []
            text_cols = []
            stats = {}
            for idx, col in enumerate(df.columns):
                if col == 'Policy Assessment':
                    continue
                if pd.api.types.is_numeric_dtype(df[col]):
                    numeric_cols.append(idx + 1)
                    abs_data = df[col].abs()
                    valid_data = abs_data[abs_data > 0]
                    
                    if len(valid_data) >= 4:
                        q1 = valid_data.quantile(0.25)
                        q3 = valid_data.quantile(0.75)
                        iqr = q3 - q1
                        if iqr == 0:
                            iqr = valid_data.mean() * 0.10
                        base_val = q3 + (1.5 * iqr)
                    else:
                        base_val = valid_data.median() + (valid_data.mean() * 0.50) if not valid_data.empty else 1.0
                    
                    if pd.isna(base_val) or base_val <= 0:
                        base_val = valid_data.mean() if not valid_data.empty and valid_data.mean() > 0 else 1.0
                        
                    stats[idx + 1] = {'base_val': base_val}
                elif pd.api.types.is_string_dtype(df[col]) or pd.api.types.is_object_dtype(df[col]):
                    text_cols.append(idx + 1)

            severity_counts = {
                "Severe": 0, "Suspicious": 0, "Slight": 0, 
                "Clean": 0, "Policy Violation": 0
            }

            for row in range(2, len(df) + 2):
                row_severity = "Clean"
                
                for col_idx in numeric_cols:
                    cell = worksheet.cell(row=row, column=col_idx)
                    val = cell.value
                    if val is not None and isinstance(val, (int, float)):
                        base_val = stats[col_idx]['base_val']
                        ratio = abs(val) / base_val
                        
                        if ratio >= 5.0:
                            cell.fill = severe_fill
                            cell.font = severe_font
                            row_severity = "Severe"
                        elif ratio >= 2.5:
                            cell.fill = red_fill
                            cell.font = red_font
                            if row_severity != "Severe":
                                row_severity = "Suspicious"
                        elif ratio >= 1.5:
                            cell.fill = yellow_fill
                            cell.font = yellow_font
                            if row_severity not in ["Severe", "Suspicious"]:
                                row_severity = "Slight"
                                
                # AI Check
                ai_cell = worksheet.cell(row=row, column=ai_col_idx)
                ai_val = str(ai_cell.value).lower()
                is_normal = "compliant" in ai_val or "missing" in ai_val or "disabled" in ai_val or "error" in ai_val
                
                if not is_normal:
                    ai_cell.fill = purple_fill
                    ai_cell.font = purple_font
                    if row_severity == "Clean" or row_severity == "Slight":
                        row_severity = "Policy Violation"
                        
                severity_counts[row_severity] += 1

            dashboard = writer.book.create_sheet("Summary Dashboard", 0)
            dashboard.sheet_view.showGridLines = False
            
            dashboard.column_dimensions['A'].width = 25
            dashboard.column_dimensions['B'].width = 15
            dashboard.column_dimensions['C'].width = 15
            dashboard.column_dimensions['D'].width = 20
            
            title_cell = dashboard['A1']
            title_cell.value = "Financial Anomaly Dashboard"
            title_cell.font = Font(size=18, bold=True, color="FF2F5597")
            
            warning_cell = dashboard['C1']
            warning_text = "(Note: May be inaccurate for small datasets)"
            if not date_col:
                warning_text += " | WARNING: No Date/Time column found. Time-based anomalies disabled."
            if global_warnings:
                warning_text += " | " + " ".join(global_warnings)
            warning_cell.value = warning_text
            warning_cell.font = Font(size=10, italic=True, color="FFFF0000")
            
            total_rows = len(df)
            total_anomalies = total_rows - severity_counts["Clean"]
            pct_anomalies = (total_anomalies / total_rows) * 100 if total_rows > 0 else 0
            
            dashboard['A3'] = "Total Transactions:"
            dashboard['B3'] = total_rows
            dashboard['A4'] = "Total Anomalies:"
            dashboard['B4'] = total_anomalies
            dashboard['A5'] = "Suspicious Rate:"
            dashboard['B5'] = f"{pct_anomalies:.1f}%"
            dashboard['A6'] = "Rapid Successions (Warnings):"
            dashboard['B6'] = rapid_warnings
            dashboard['A7'] = "Policy Violations:"
            dashboard['B7'] = severity_counts["Policy Violation"]

            for r in range(3, 8):
                dashboard[f'A{r}'].font = Font(bold=True)
                dashboard[f'B{r}'].alignment = Alignment(horizontal='left')
                if r in [4, 5]:
                    color = "FF990000" if total_anomalies > 0 else "FF006100"
                    dashboard[f'B{r}'].font = Font(bold=True, color=color)
                elif r == 6:
                    color = "FF9C0006" if rapid_warnings > 0 else "FF006100"
                    dashboard[f'B{r}'].font = Font(bold=True, color=color)
                elif r == 7:
                    color = "FF333399" if severity_counts["Policy Violation"] > 0 else "FF006100"
                    dashboard[f'B{r}'].font = Font(bold=True, color=color)
                else:
                    dashboard[f'B{r}'].font = Font(bold=True)
                    
            dashboard['A9'] = "Anomaly Type"
            dashboard['B9'] = "Count"
            dashboard['C9'] = "% of Total"
            dashboard['D9'] = "% of Anomalies"
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for col in ['A', 'B', 'C', 'D']:
                dashboard[f'{col}9'].font = Font(bold=True, color="FFFFFFFF")
                dashboard[f'{col}9'].fill = PatternFill(start_color="FF2F5597", end_color="FF2F5597", fill_type="solid")
                dashboard[f'{col}9'].alignment = Alignment(horizontal="center", vertical="center")
                dashboard[f'{col}9'].border = thin_border
                
            categories = ["Severe", "Suspicious", "Slight", "Policy Violation", "Clean"]
            for i, cat in enumerate(categories):
                row = 10 + i
                count = severity_counts[cat]
                pct_total = (count / total_rows) if total_rows > 0 else 0
                pct_anom = (count / total_anomalies) if (total_anomalies > 0 and cat != "Clean") else 0
                
                dashboard[f'A{row}'] = cat
                dashboard[f'B{row}'] = count
                dashboard[f'C{row}'] = pct_total
                dashboard[f'C{row}'].number_format = '0.0%'
                
                if cat == "Clean":
                    dashboard[f'D{row}'] = "N/A"
                else:
                    dashboard[f'D{row}'] = pct_anom
                    dashboard[f'D{row}'].number_format = '0.0%'
                    
                for col in ['A', 'B', 'C', 'D']:
                    cell = dashboard[f'{col}{row}']
                    cell.border = thin_border
                    if col in ['B', 'C', 'D']:
                        cell.alignment = Alignment(horizontal="center")
                    
            pie1 = PieChart3D()
            pie1.title = "Dataset Breakdown (Whole Set)"
            labels1 = Reference(dashboard, min_col=1, min_row=10, max_row=15)
            data1 = Reference(dashboard, min_col=2, min_row=9, max_row=15)
            pie1.add_data(data1, titles_from_data=True)
            pie1.set_categories(labels1)
            pie1.dataLabels = DataLabelList()
            pie1.dataLabels.showSerName = False
            pie1.dataLabels.showVal = False
            pie1.dataLabels.showCatName = True
            pie1.dataLabels.showPercent = True
            pie1.legend = None
            pie1.width = 12
            pie1.height = 8
            dashboard.add_chart(pie1, "F2")
            
            if total_anomalies > 0:
                pie2 = PieChart3D()
                pie2.title = "Anomaly Distribution (Anomalies Only)"
                labels2 = Reference(dashboard, min_col=1, min_row=10, max_row=14)
                data2 = Reference(dashboard, min_col=2, min_row=9, max_row=14)
                pie2.add_data(data2, titles_from_data=True)
                pie2.set_categories(labels2)
                pie2.dataLabels = DataLabelList()
                pie2.dataLabels.showSerName = False
                pie2.dataLabels.showVal = False
                pie2.dataLabels.showCatName = True
                pie2.dataLabels.showPercent = True
                pie2.legend = None
                pie2.width = 12
                pie2.height = 8
                dashboard.add_chart(pie2, "F17")
                    
            writer.book.active = 0
            
        TASK_STORE[task_id]["message"] = "Completed!"
        TASK_STORE[task_id]["progress"] = 100
        TASK_STORE[task_id]["result"] = buffer.getvalue()
        TASK_STORE[task_id]["status"] = "completed"

    except Exception as e:
        print(f"Error in background task: {e}")
        TASK_STORE[task_id]["status"] = "error"
        TASK_STORE[task_id]["message"] = f"Error: {str(e)}"


import secrets
import urllib.request
import httpx
import secrets
import json

class ForgotPasswordRequest(BaseModel):
    email: str

@app.post("/forgot-password")
def forgot_password(req: ForgotPasswordRequest, db: Session = Depends(get_db)):
    email_lower = req.email.lower()
    user = db.query(User).filter(User.email == email_lower).first()
    if not user:
        return {"message": "If that account exists, a recovery email has been sent."}
    
    token = secrets.token_urlsafe(32)
    user.reset_token = token
    user.reset_token_expiry = datetime.utcnow() + timedelta(minutes=15)
    db.commit()
    
    reset_link = f"https://financial-data-parser.onrender.com/reset_password.html?token={token}"
    
    # Send email via Resend API
    resend_api_key = os.getenv("RESEND_API_KEY")
    if resend_api_key:
        try:
            req_data = json.dumps({
                "from": "onboarding@resend.dev",
                "to": [user.email],
                "subject": "Password Reset Request",
                "html": f"<p>Click the link below to reset your password:</p><p><a href='{reset_link}'>{reset_link}</a></p><p>This link expires in 15 minutes.</p>"
            }).encode('utf-8')
            
            headers = {
                "Authorization": f"Bearer {resend_api_key}",
                "Content-Type": "application/json"
            }
            
            http_req = urllib.request.Request("https://api.resend.com/emails", data=req_data, headers=headers, method="POST")
            urllib.request.urlopen(http_req)
        except Exception as e:
            print("Failed to send email:", e)
    else:
        print(f"\n[MOCK EMAIL] To: {user.email}\nSubject: Password Reset\nLink: {reset_link}\n")
        
    return {"message": "If that account exists, a recovery email has been sent."}

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

@app.post("/reset-password")
def reset_password(req: ResetPasswordRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.reset_token == req.token).first()
    if not user or user.reset_token_expiry < datetime.utcnow():
        raise HTTPException(status_code=400, detail="Invalid or expired token")
        
    user.hashed_password = get_password_hash(req.new_password)
    user.reset_token = None
    user.reset_token_expiry = None
    # Also unlock the account if it was locked!
    user.failed_login_attempts = 0
    user.lockout_count = 0
    user.locked_until = None
    db.commit()
    
    return {"message": "Password has been reset successfully!"}


import stripe
from fastapi import Request

@app.get("/user/me")
def get_user_me(current_user: User = Depends(get_current_user)):
    tier = current_user.subscription_tier or 'free'
    if tier == 'free':
        limit = 100
    elif tier == 'budget':
        limit = 1000
    elif tier == 'pro':
        limit = 5000
    else:
        limit = 'Unlimited'
        
    now = datetime.now()
    if not current_user.last_reset_date or current_user.last_reset_date.month != now.month or current_user.last_reset_date.year != now.year:
        usage = 0
    else:
        usage = current_user.rows_processed_this_month
        
    return {
        "email": current_user.email,
        "rows_processed_this_month": usage,
        "limit": limit,
        "plan": tier
    }

class CheckoutRequest(BaseModel):
    package: str # "basic" or "pro"

@app.post("/create-checkout-session")
def create_checkout_session(req: CheckoutRequest, current_user: User = Depends(get_current_user)):
    stripe.api_key = os.getenv("STRIPE_SECRET_KEY")
    if not stripe.api_key:
        raise HTTPException(status_code=500, detail="Stripe API keys are not configured on this server.")
        
    prices = {
        "budget": {"price": 500, "name": "Budget Tier (1,000 Rows/mo)"},
        "pro": {"price": 1500, "name": "Pro Tier (5,000 Rows/mo)"},
        "unlimited": {"price": 2900, "name": "Unlimited Tier"}
    }
    
    if req.package not in prices:
        raise HTTPException(status_code=400, detail="Invalid package selected.")
        
    pkg = prices[req.package]
    
    try:
        session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': 'usd',
                    'product_data': {
                        'name': pkg['name'],
                        'description': 'Monthly subscription for FinParse.',
                    },
                    'unit_amount': pkg['price'],
                    'recurring': {'interval': 'month'},
                },
                'quantity': 1,
            }],
            mode='subscription',
            success_url="https://financial-data-parser.onrender.com/dashboard.html?payment=success",
            cancel_url="https://financial-data-parser.onrender.com/pricing.html?payment=cancelled",
            client_reference_id=str(current_user.id),
            metadata={"tier": req.package}
        )
        return {"checkout_url": session.url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/stripe/webhook")
async def stripe_webhook(request: Request, db: Session = Depends(get_db)):
    stripe.api_key = os.getenv("STRIPE_SECRET_KEY")
    webhook_secret = os.getenv("STRIPE_WEBHOOK_SECRET")
    
    if not stripe.api_key or not webhook_secret:
        return {"status": "Webhook ignored (keys not configured)"}

    payload = await request.body()
    sig_header = request.headers.get("stripe-signature")

    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, webhook_secret
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Invalid payload")
    except stripe.error.SignatureVerificationError as e:
        raise HTTPException(status_code=400, detail="Invalid signature")

    if event['type'] == 'checkout.session.completed' or event['type'] == 'invoice.payment_succeeded':
        # Every month when the invoice is paid, we reset the user's credits to their plan limit
        event_id = event['id']
        if db.query(ProcessedEvent).filter(ProcessedEvent.id == event_id).first():
            return {"status": "already processed"}
            
        obj = event['data']['object']
        
        # When checkout completes, we get client_reference_id and metadata directly
        if event['type'] == 'checkout.session.completed':
            user_id = obj.get("client_reference_id")
            tier = obj.get("metadata", {}).get("tier")
            
            if user_id and tier:
                user = db.query(User).filter(User.id == int(user_id)).first()
                if user:
                    user.subscription_tier = tier
                    user.rows_processed_this_month = 0
                    user.last_reset_date = datetime.now()
                    db.add(ProcessedEvent(id=event_id))
                    db.commit()
                    print(f"Activated {tier} plan for User {user.id}")
        
        # NOTE: For invoice.payment_succeeded (monthly recurring), 
        # finding the user and metadata requires hitting the Stripe API for the Subscription object.
        # For this MVP, we only process the initial checkout session. 
        # To make recurring work fully, you would query `stripe.Subscription.retrieve(obj.subscription)`.

    return {"status": "success"}


@app.get("/reset_db")
def reset_db():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    return {"status": "Database has been completely reset to the new schema!"}

# --- GOOGLE OAUTH ---
import secrets
import httpx
from fastapi.responses import RedirectResponse

GOOGLE_CLIENT_ID = os.getenv('GOOGLE_CLIENT_ID')
GOOGLE_CLIENT_SECRET = os.getenv('GOOGLE_CLIENT_SECRET')
GOOGLE_REDIRECT_URI = os.getenv('GOOGLE_REDIRECT_URI', 'https://financial-data-parser.onrender.com/auth/google/callback')

@app.get('/login/google')
def login_google():
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=500, detail='Google Client ID not configured.')
    
    auth_url = (
        f'https://accounts.google.com/o/oauth2/v2/auth?'
        f'client_id={GOOGLE_CLIENT_ID}&'
        f'redirect_uri={GOOGLE_REDIRECT_URI}&'
        f'response_type=code&'
        f'scope=openid%20email%20profile&'
        f'access_type=offline'
    )
    return RedirectResponse(url=auth_url)

@app.get('/auth/google/callback')
async def auth_google_callback(code: str, db: Session = Depends(get_db)):
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        raise HTTPException(status_code=500, detail='Google credentials not configured.')
        
    token_url = 'https://oauth2.googleapis.com/token'
    data = {
        'code': code,
        'client_id': GOOGLE_CLIENT_ID,
        'client_secret': GOOGLE_CLIENT_SECRET,
        'redirect_uri': GOOGLE_REDIRECT_URI,
        'grant_type': 'authorization_code',
    }
    
    async with httpx.AsyncClient() as client:
        # 1. Exchange code for token
        response = await client.post(token_url, data=data)
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail=f'Failed to fetch token: {response.text}')
        
        token_data = response.json()
        access_token = token_data.get('access_token')
        
        # 2. Fetch user info
        user_info_url = 'https://www.googleapis.com/oauth2/v2/userinfo'
        user_res = await client.get(user_info_url, headers={'Authorization': f'Bearer {access_token}'})
        if user_res.status_code != 200:
            raise HTTPException(status_code=400, detail='Failed to fetch user info')
            
        user_info = user_res.json()
        email_raw = user_info.get('email')
        
        if not email_raw:
            raise HTTPException(status_code=400, detail='Email not provided by Google')
            
        email_lower = email_raw.lower()
            
        # 3. Check if user exists, if not create them
        user = db.query(User).filter(User.email == email_lower).first()
        if not user:
            # Create account with random password
            random_pwd = secrets.token_urlsafe(16)
            hashed_pwd = get_password_hash(random_pwd)
            tier = 'unlimited' if email_lower == 'mg.shuanchi@gmail.com' else 'free'
            user = User(email=email_lower, hashed_password=hashed_pwd, subscription_tier=tier, is_verified=True)
            db.add(user)
            db.commit()
            db.refresh(user)
        else:
            if email_lower == 'mg.shuanchi@gmail.com' and user.subscription_tier != 'unlimited':
                user.subscription_tier = 'unlimited'
                db.commit()
            
        # 4. Generate our JWT Token
        jwt_token = create_access_token(data={'sub': user.email})
        
        # 5. Redirect back to frontend dashboard with token in URL (frontend JS will catch and save it)
        return RedirectResponse(url=f'https://financial-data-parser.onrender.com/dashboard.html?token={jwt_token}')

import os
if os.path.exists("frontend"):
    app.mount("/", StaticFiles(directory="frontend", html=True), name="frontend")
elif os.path.exists("../frontend"):
    app.mount("/", StaticFiles(directory="../frontend", html=True), name="frontend")
